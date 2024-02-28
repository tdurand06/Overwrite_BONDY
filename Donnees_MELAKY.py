import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
#Variables
cell_width = 20
#Constant
# Explicit list of village names to keep
hameau_to_keep = [
    "Vakivao", "Ambalahonko", "Amboanio", "Tsiamandira", "Ankorovahiny",
    "Ambato"
]
# The path to your original Excel file
original_file_path = '/Users/thomasdurand/Desktop/BONDY/diana_excel/données_brutes_MELAKY.xlsx'
# The path for the new modified Excel file
new_file_path = '/Users/thomasdurand/Desktop/BONDY/diana_excel/données_filtrées_MELAKY.xlsx'
# Tabs to update based on the "Informations générales" filtering, excluding "Informations générales"
tabs_to_update = [
    "Revenues", "Capital", "Agriculture", "Elevage", "Apiculture", "Pêche",
    "Exploitation forestière", "Dépenses", "Educ_Santé", "Place_femme"
]

# Load the "Informations générales" tab to get the IDs and Fokontany to keep
df_info_gen = pd.read_excel(original_file_path, sheet_name='BD_Enquête_P1')
IDs_to_keep = df_info_gen[df_info_gen['Hameau'].isin(hameau_to_keep)]['ID_Ménage'].tolist()
# Create a mapping of ID_Ménage to Fokontany for later use
df_info_gen_unique = df_info_gen.drop_duplicates(subset='ID_Ménage', keep='first')
id_to_hameau = df_info_gen_unique.set_index('ID_Ménage')['Hameau'].to_dict()

# Initialize Excel writer
with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
    # Save "Informations générales" tab as is
    df_info_gen_filtered = df_info_gen[df_info_gen['ID_Ménage'].isin(IDs_to_keep)]
    df_info_gen_filtered = df_info_gen_filtered.sort_values(by='Hameau', ignore_index=True)
    df_info_gen_filtered.to_excel(writer, sheet_name='Informations générales', index=False)

    # Process each specified tab
    for tab in tabs_to_update:
        if tab in pd.ExcelFile(original_file_path).sheet_names:
            df_tab = pd.read_excel(original_file_path, sheet_name=tab)
            # Filter rows based on the IDs_to_keep
            df_filtered = df_tab[df_tab['ID_Ménage'].isin(IDs_to_keep)]
            # Add Fokontany column from mapping
            df_filtered.insert(1, 'Hameau', df_filtered['ID_Ménage'].map(id_to_hameau))
            df_filtered = df_filtered.sort_values(by='Hameau', ignore_index=True)
            df_filtered.to_excel(writer, sheet_name=tab, index=False)

print("The modified file has been saved with the 'Fokontany' column added to the specified tabs.")

wb = load_workbook(new_file_path)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(wrapText=True)
        ws.column_dimensions[col[0].column_letter].width = cell_width
wb.save(new_file_path)
print("The cell should be adjusted")
#This should work

